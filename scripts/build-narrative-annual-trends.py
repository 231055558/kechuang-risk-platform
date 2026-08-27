#!/usr/bin/env python3
"""Build the sanitized revised-method annual narrative-risk snapshot.

The script reads privately archived public filings and transcripts. It never
copies full text or local paths into the public snapshot.
"""

from __future__ import annotations

import hashlib
import html
import json
import math
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PRIVATE_ROOT = ROOT / "private/narrative-risk/annual"
PYTHON_DEPS = PRIVATE_ROOT / "python-deps"
if PYTHON_DEPS.exists():
    sys.path.insert(0, str(PYTHON_DEPS))

import jieba  # type: ignore  # installed into the private task runtime
from openpyxl import load_workbook
from pypdf import PdfReader

METHOD_VERSION = "narrative-method-revised-2026-08-27-v2"
AS_OF_DATE = "2026-08-27"
SCHEMA_VERSION = "KCR-NARRATIVE-RISK-2026.08-v1"
REPORT_MANIFEST_PATH = ROOT / "config/narrative-risk-annual-reports.json"
STOPWORDS_PATH = ROOT / "config/narrative-risk-stopwords.txt"
INNOVATION_LEXICON_PATH = (
    PRIVATE_ROOT / "dictionaries/程新生论文2022创新信息披露关键词词典.xlsx"
)
TONE_MANIFEST_PATH = ROOT / "config/narrative-risk-tone-sources.json"
METHOD_DOCUMENT_PATH = Path("/Users/mao/Downloads/叙事风险维度测度（修改版）.docx")
SENTIMENT_DICTIONARY_PATH = (
    PRIVATE_ROOT / "dictionaries/中文金融情感词典_姜富伟等(2020).xlsx"
)
SNAPSHOT_PATH = ROOT / "src/data/industry/narrative-risk-annual-trends.json"
EXTRACTED_TEXT_ROOT = PRIVATE_ROOT / "extracted-text"

UNCERTAINTY_WORDS = [
    "风险",
    "可能",
    "不确定性",
    "潜在",
    "或",
    "有望",
    "预计",
    "一定程度上",
]

# Only observations with an annual invention-application count explicitly
# disclosed in the corresponding official annual report are entered. The
# disclosure basis is retained in the sanitized observation details.
ANNUAL_REPORT_INVENTION_APPLICATIONS = {
    ("cambricon", 2021): (380, "年报知识产权表：发明专利本年新增申请数"),
    ("cambricon", 2022): (184, "年报研发成果：新增发明专利申请数"),
    ("cambricon", 2023): (113, "年报研发成果：新增发明专利申请数"),
    ("cambricon", 2024): (102, "年报研发成果：新增发明专利申请数"),
    ("cambricon", 2025): (101, "年报研发成果：新增发明专利申请数"),
    ("hengrui-pharma", 2021): (244, "年报研发成果：国内新申请专利数"),
    ("hengrui-pharma", 2022): (169, "年报研发成果：国内新申请专利数"),
    ("hengrui-pharma", 2023): (246, "年报研发成果：国内新申请专利数"),
    ("hengrui-pharma", 2024): (456, "年报研发成果：大中华地区新申请专利数"),
    ("hengrui-pharma", 2025): (459, "年报研发成果：大中华地区新申请专利数"),
    ("baili-tianheng", 2023): (109, "年报知识产权表：发明专利本年新增申请数"),
    ("baili-tianheng", 2024): (159, "年报知识产权表：发明专利本年新增申请数"),
    ("baili-tianheng", 2025): (173, "年报知识产权表：发明专利本年新增申请数"),
    ("huami-electronics", 2024): (16, "年报研发成果：申请发明专利数"),
    ("huami-electronics", 2025): (13, "年报研发成果：申请发明专利数"),
}

COMPANY_META = {
    "cambricon": ("寒武纪", "688256", "上市", [2021, 2022, 2023, 2024, 2025]),
    "hengrui-pharma": ("恒瑞医药", "600276", "上市", [2021, 2022, 2023, 2024, 2025]),
    "huami-electronics": ("华微电子", "600360", "上市", [2021, 2022, 2023, 2024, 2025]),
    "baili-tianheng": ("百利天恒", "688506", "上市", [2023, 2024, 2025]),
    "zuojiang-technology": ("左江科技", "300799", "已退市", [2021, 2022, 2023]),
}

METRICS = [
    {
        "metricKey": "information_sufficiency",
        "name": "信息总量充分性",
        "category": "年报披露质量",
        "formula": "经营情况讨论与分析及风险章节的有效词数÷10,000",
        "unit": "万有效词",
        "riskDirection": "数值越高，信息容量越充分；过长文本仍需结合其他指标判断",
        "methodStatus": "新版正式口径",
    },
    {
        "metricKey": "risk_context_ambiguity",
        "name": "风险语境模糊度",
        "category": "年报披露质量",
        "formula": "业绩展望及风险提示中的模糊词次数÷该范围有效词数",
        "unit": "比例",
        "riskDirection": "数值越高，风险叙述越模糊",
        "methodStatus": "新版正式口径",
    },
    {
        "metricKey": "data_support_ratio",
        "name": "信息数据支撑度",
        "category": "年报披露质量",
        "formula": "定量信息出现次数÷同一披露范围有效词数",
        "unit": "比例",
        "riskDirection": "数值越高，叙述的定量支撑越充分",
        "methodStatus": "新版正式口径",
    },
    {
        "metricKey": "disclosure_quality",
        "name": "年报披露质量",
        "category": "年报披露质量",
        "formula": "信息充分性归一值、风险语境清晰度和数据支撑度三项等权平均",
        "unit": "指数",
        "riskDirection": "数值越高，披露质量越高、叙事风险越低",
        "methodStatus": "新版解释确认口径",
    },
    {
        "metricKey": "innovation_talk_density",
        "name": "创新文本密度",
        "category": "叙事夸大性",
        "formula": "管理层讨论与分析及公司业务概要中的创新词次数÷有效词数×1,000",
        "unit": "次/千词",
        "riskDirection": "数值越高表示创新叙事投入越强，需结合创新行动与叙事夸大度判断",
        "methodStatus": "程新生等（2022）完整词典口径",
    },
    {
        "metricKey": "innovation_action_strength",
        "name": "创新行动强度",
        "category": "叙事夸大性",
        "formula": "当年发明专利申请数加一后取自然对数",
        "unit": "对数值",
        "riskDirection": "数值越高，已公开验证的创新行动越强",
        "methodStatus": "年报明确披露值；专利明细待国家知识产权局逐项复核",
    },
    {
        "metricKey": "innovation_divergence",
        "name": "叙事夸大度",
        "category": "叙事夸大性",
        "formula": "创新文本密度加一取自然对数，减去当年发明专利申请数加一的自然对数",
        "unit": "对数差",
        "riskDirection": "大于零为多言寡行；小于或等于零为言行匹配或寡言多行",
        "methodStatus": "新版方案二：企业自身言行对数差",
    },
    {
        "metricKey": "positive_tone_intensity",
        "name": "正面语调强度",
        "category": "管理者语调",
        "formula": "管理层回答中的正面情感词数÷管理层回答有效词数",
        "unit": "比例",
        "riskDirection": "异常高值需结合负面语调和净正面语调判断",
        "methodStatus": "新版正式词频口径",
    },
    {
        "metricKey": "negative_tone_intensity",
        "name": "负面语调强度",
        "category": "管理者语调",
        "formula": "管理层回答中的负面情感词数÷管理层回答有效词数",
        "unit": "比例",
        "riskDirection": "数值越高，负面风险信号越强",
        "methodStatus": "新版正式词频口径",
    },
    {
        "metricKey": "manager_net_positive_tone",
        "name": "管理者净正面语调",
        "category": "管理者语调",
        "formula": "正面语调强度减负面语调强度，再除以两者之和",
        "unit": "指数",
        "riskDirection": "负值为高风险，0至0.4为中风险，0.4至1为低风险",
        "methodStatus": "新版正式词频口径",
    },
]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def load_word_list(path: Path) -> list[str]:
    return [line.strip() for line in path.read_text("utf-8").splitlines() if line.strip()]


def load_innovation_dictionary(path: Path) -> tuple[list[str], Counter]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["逐词词典"]
    terms: list[str] = []
    categories: Counter = Counter()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        category, _, term = row[:3]
        if isinstance(category, str) and isinstance(term, str) and term.strip():
            terms.append(term.strip())
            categories[category.strip()] += 1
    return terms, categories


def expand_innovation_term(term: str) -> str:
    pattern = re.escape(term)
    pattern = pattern.replace(re.escape("（的）"), "(?:的)?")
    for roman, chinese in (("Ⅰ", "一"), ("Ⅱ", "二"), ("Ⅲ", "三"), ("Ⅳ", "四")):
        pattern = pattern.replace(
            re.escape(f"{roman}（{chinese}）"),
            f"(?:{re.escape(roman)}|{re.escape(chinese)})",
        )
    return pattern


STOPWORDS = set(load_word_list(STOPWORDS_PATH))
INNOVATION_DICTIONARY_ROWS, INNOVATION_CATEGORY_COUNTS = load_innovation_dictionary(
    INNOVATION_LEXICON_PATH
)
INNOVATION_TERMS = sorted(set(INNOVATION_DICTIONARY_ROWS))
INNOVATION_PATTERN = re.compile(
    "|".join(
        expand_innovation_term(term)
        for term in sorted(INNOVATION_TERMS, key=lambda item: (-len(item), item))
    )
)


def clean_pdf_text(value: str) -> str:
    value = value.replace("\u00a0", " ").replace("\x00", "")
    value = re.sub(r"[^\n]*?\d{4}\s*年年度报告\s*\d+\s*/\s*\d+[^\n]*", "\n", value)
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


SECTION_HEADER = re.compile(r"(?:^|\n)\s*第[一二三四五六七八九十百]+节\s*", re.M)


def main_section(text: str, titles: list[str]) -> tuple[str, str | None]:
    candidates: list[tuple[int, int, str]] = []
    title_group = "|".join(re.escape(title) for title in titles)
    pattern = re.compile(
        rf"(?:^|\n)\s*第[一二三四五六七八九十百]+节\s*(?:{title_group})\s*(?:\n|$)",
        re.M,
    )
    for match in pattern.finditer(text):
        next_header = SECTION_HEADER.search(text, match.end())
        end = next_header.start() if next_header else len(text)
        candidates.append((match.start(), end, match.group(0).strip()))
    if not candidates:
        return "", None
    start, end, heading = max(candidates, key=lambda item: item[1] - item[0])
    section = text[start:end].strip()
    if len(section) < 500:
        return "", None
    return section, heading


def numbered_subsection(text: str, titles: list[str]) -> tuple[str, str | None]:
    title_group = "|".join(re.escape(title) for title in titles)
    pattern = re.compile(
        rf"(?:^|\n)\s*(?:(?P<top>[一二三四五六七八九十百]+)\s*、|[（(](?P<paren>[一二三四五六七八九十百]+)[）)]|(?P<digit>\d+)\s*[、.．])\s*(?:{title_group})[^\n]*",
        re.M,
    )
    matches = list(pattern.finditer(text))
    if not matches:
        plain = re.compile(rf"(?:^|\n)\s*(?:{title_group})\s*(?:\n|$)", re.M)
        matches = list(plain.finditer(text))
    if not matches:
        return "", None
    candidates = []
    for match in matches:
        if match.groupdict().get("top"):
            next_heading_pattern = re.compile(
                r"(?:^|\n)\s*[一二三四五六七八九十百]+\s*、\s*[^\n]{2,80}", re.M
            )
        elif match.groupdict().get("paren"):
            next_heading_pattern = re.compile(
                r"(?:^|\n)\s*[（(][一二三四五六七八九十百]+[）)]\s*[^\n]{2,80}", re.M
            )
        else:
            next_heading_pattern = re.compile(
                r"(?:^|\n)\s*\d+\s*[、.．]\s*[^\n]{2,80}", re.M
            )
        next_heading = next_heading_pattern.search(text, match.end())
        end = next_heading.start() if next_heading else len(text)
        candidates.append((match.start(), end, match.group(0).strip()))
    start, end, heading = max(candidates, key=lambda item: item[1] - item[0])
    section = text[start:end].strip()
    return (section, heading) if len(section) >= 80 else ("", None)


def combine_non_overlapping(*sections: str) -> str:
    selected: list[str] = []
    for section in sections:
        if not section:
            continue
        if any(section in existing for existing in selected):
            continue
        selected = [existing for existing in selected if existing not in section]
        selected.append(section)
    return "\n".join(selected)


def tokenize(value: str, *, remove_numeric: bool = False) -> list[str]:
    tokens: list[str] = []
    for raw in jieba.lcut(value, cut_all=False):
        token = raw.strip().lower()
        if not token or token in STOPWORDS:
            continue
        if remove_numeric and re.fullmatch(r"[\d.,%％+-]+", token):
            continue
        if re.search(r"[\u3400-\u9fff]", token) or (not remove_numeric and re.search(r"\d", token)):
            tokens.append(token)
    return tokens


NUMERIC_PATTERN = re.compile(
    r"(?<![A-Za-z])[-+]?\d[\d,]*(?:\.\d+)?\s*(?:%|％|亿元|万元|元|倍|个百分点)?"
)
FINANCIAL_RATIO_PATTERN = re.compile(
    r"(?:毛利率|净利率|资产负债率|流动比率|速动比率|研发投入占比|同比|环比)"
)


def innovation_hits(value: str) -> int:
    return sum(1 for _ in INNOVATION_PATTERN.finditer(value))


def publication_date(url: str) -> str | None:
    match = re.search(r"/(20\d{2})-(\d{2})-(\d{2})/", url)
    return "-".join(match.groups()) if match else None


def risk_label(value: float | None) -> str | None:
    if value is None:
        return None
    if value < 0:
        return "高风险"
    if value < 0.4:
        return "中风险"
    return "低风险"


def clamp(value: float, lower: float = 0.0, upper: float = 1.0) -> float:
    return min(upper, max(lower, value))


def apply_risk_scores(observations: list[dict]) -> None:
    values_by_metric = {
        metric["metricKey"]: [
            float(item["value"])
            for item in observations
            if item["metricKey"] == metric["metricKey"] and item["value"] is not None
        ]
        for metric in METRICS
    }
    higher_is_riskier_metrics = {
        "risk_context_ambiguity",
        "innovation_talk_density",
        "innovation_divergence",
        "positive_tone_intensity",
        "negative_tone_intensity",
    }
    mapping_by_metric = {}
    for metric in METRICS:
        metric_key = metric["metricKey"]
        values = values_by_metric[metric_key]
        minimum = min(values)
        maximum = max(values)
        higher_is_riskier = metric_key in higher_is_riskier_metrics
        scorer = lambda value, minimum=minimum, maximum=maximum, higher_is_riskier=higher_is_riskier: (
            50.0
            if maximum == minimum
            else (
                clamp((value - minimum) / (maximum - minimum))
                if higher_is_riskier
                else 1 - clamp((value - minimum) / (maximum - minimum))
            )
            * 100
        )
        parameters = [
            {"name": "有效观测数", "value": str(len(values))},
            {"name": "样本最小值", "value": f"{minimum:.10g}"},
            {"name": "样本最大值", "value": f"{maximum:.10g}"},
        ]
        if metric_key == "information_sufficiency":
            parameters.append({"name": "方法参考区间", "value": "0.5至1.0万有效词"})
        elif metric_key == "innovation_divergence":
            parameters.append({"name": "多言寡行原始分界", "value": "0"})
        elif metric_key == "manager_net_positive_tone":
            parameters.append({"name": "原始风险区间", "value": "负值高；0至0.4中；0.4至1低"})
        mapping = {
            "name": "当前样本极差" + ("正向" if higher_is_riskier else "反向") + "映射",
            "formula": (
                "(原始值−样本最小值)÷(样本最大值−样本最小值)×100"
                if higher_is_riskier
                else "(样本最大值−原始值)÷(样本最大值−样本最小值)×100"
            ),
            "parameterSource": "最新版方法文件确定风险方向；样本参数取五家上市企业2021—2025窗口内该指标的全部有效原始观测",
            "parameters": parameters,
            "limitation": "0和100表示当前五家上市样本窗口内的相对最低与相对最高风险；样本或窗口改变后需重算，不代表跨样本绝对风险。",
        }

        metric["riskMapping"] = mapping
        mapping_by_metric[metric_key] = (scorer, mapping)

    for item in observations:
        if item["value"] is None:
            item["riskScore"] = None
            continue
        scorer, _ = mapping_by_metric[item["metricKey"]]
        item["riskScore"] = round(clamp(float(scorer(float(item["value"]))), 0, 100), 10)

    observation_index = {
        (item["companyKey"], item["year"], item["metricKey"]): item
        for item in observations
    }
    for company_key, (_, _, _, years) in COMPANY_META.items():
        for metric in METRICS:
            previous = None
            for year in years:
                item = observation_index.get((company_key, year, metric["metricKey"]))
                if item is None:
                    continue
                current = item["riskScore"]
                item["riskScoreChange"] = (
                    round(current - previous, 10)
                    if current is not None and previous is not None
                    else None
                )
                previous = current


def value_record(
    company_key: str,
    year: int,
    metric_key: str,
    value: float | None,
    status: str,
    reason: str | None,
    document_id: str | None,
    details: dict | None = None,
) -> dict:
    return {
        "companyKey": company_key,
        "year": year,
        "metricKey": metric_key,
        "value": round(value, 10) if value is not None else None,
        "changeRate": None,
        "riskScore": None,
        "riskScoreChange": None,
        "status": status,
        "missingReason": reason,
        "documentId": document_id,
        "methodVersion": METHOD_VERSION,
        "details": details or {},
    }


def load_sentiment_dictionary() -> tuple[set[str], set[str], str]:
    workbook = load_workbook(SENTIMENT_DICTIONARY_PATH, read_only=True, data_only=True)
    result = {}
    for sheet_name in ("positive", "negative"):
        words = set()
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if isinstance(row[0], str) and row[0].strip():
                words.add(row[0].strip())
        result[sheet_name] = words
    return result["positive"], result["negative"], sha256_file(SENTIMENT_DICTIONARY_PATH)


def clean_answer_text(value: str) -> str:
    value = html.unescape(value)
    value = re.sub(r"<br\s*/?>", "\n", value, flags=re.I)
    value = re.sub(r"<[^>]+>", " ", value)
    value = re.sub(r"[\t\r ]+", " ", value)
    return value.strip()


def main() -> None:
    manifest = json.loads(REPORT_MANIFEST_PATH.read_text("utf-8"))
    tone_manifest = json.loads(TONE_MANIFEST_PATH.read_text("utf-8"))
    tone_sources = {
        (item["companyKey"], item["year"]): {
            **item,
            "url": f"https://roadshow.sseinfo.com/activityDetails/{item['activityId']}?navId=48",
            "file": PRIVATE_ROOT
            / "tone"
            / f"{item['companyKey']}-{item['year']}-{item['activityId']}.json",
        }
        for item in tone_manifest["sources"]
    }
    tone_exclusions = {
        (item["companyKey"], item["year"]): item["reason"]
        for item in tone_manifest.get("exclusions", [])
    }
    EXTRACTED_TEXT_ROOT.mkdir(parents=True, exist_ok=True)
    positive_words, negative_words, sentiment_hash = load_sentiment_dictionary()
    for word in positive_words | negative_words | set(INNOVATION_TERMS):
        jieba.add_word(word)

    companies = []
    for company_key, (name, stock_code, listing_status, years) in COMPANY_META.items():
        companies.append(
            {
                "companyKey": company_key,
                "companyName": name,
                "stockCode": stock_code,
                "listingStatus": listing_status,
                "includedYears": years,
                "exclusionReason": "未上市，不纳入年报趋势" if not years else None,
            }
        )

    documents = []
    observations = []
    report_lookup = {(r["companyKey"], r["year"]): r for r in manifest["reports"]}

    for (company_key, year), item in sorted(report_lookup.items()):
        pdf_path = PRIVATE_ROOT / "reports" / company_key / f"{year}.pdf"
        document_id = f"annual-report:{company_key}:{year}:{METHOD_VERSION}"
        if not pdf_path.exists():
            documents.append(
                {
                    "documentId": document_id,
                    "companyKey": company_key,
                    "year": year,
                    "title": f"{item['companyName']}{year}年年度报告",
                    "officialUrl": item["officialUrl"],
                    "publicationDate": publication_date(item["officialUrl"]),
                    "archiveStatus": "无法访问",
                    "parseStatus": "未解析",
                    "fileSha256": None,
                    "byteSize": None,
                    "pageCount": None,
                    "sectionCoverage": {},
                    "browserValidation": "Ego已确认公开入口；文件归档失败",
                }
            )
            for metric in METRICS[:7]:
                observations.append(
                    value_record(company_key, year, metric["metricKey"], None, "缺失", "年度报告未归档", document_id)
                )
            continue

        reader = PdfReader(str(pdf_path))
        raw_text = "\n".join((page.extract_text() or "") for page in reader.pages)
        text = clean_pdf_text(raw_text)
        text_dir = EXTRACTED_TEXT_ROOT / company_key
        text_dir.mkdir(parents=True, exist_ok=True)
        (text_dir / f"{year}.txt").write_text(text, "utf-8")

        mdna, mdna_heading = main_section(
            text,
            ["管理层讨论与分析", "经营情况讨论与分析", "董事会报告"],
        )
        business, business_heading = main_section(text, ["公司业务概要", "公司业务概述"])
        if not business and mdna:
            business, business_heading = numbered_subsection(
                mdna,
                [
                    "报告期内公司所从事的主要业务、经营模式、行业情况及研发情况说明",
                    "报告期内公司从事的主要业务",
                    "公司主要业务",
                    "主营业务分析",
                ],
            )
        risk, risk_heading = numbered_subsection(
            mdna or text,
            ["风险因素", "可能面对的风险", "公司面临的风险和应对措施", "主要风险"],
        )
        outlook, outlook_heading = numbered_subsection(
            mdna or text,
            [
                "公司关于公司未来发展的讨论与分析",
                "公司未来发展的展望",
                "未来发展展望",
                "经营计划",
                "发展战略",
            ],
        )

        disclosure_text = combine_non_overlapping(mdna, risk)
        ambiguity_text = combine_non_overlapping(outlook, risk)
        innovation_text = combine_non_overlapping(mdna, business)
        disclosure_words = tokenize(disclosure_text) if disclosure_text else []
        ambiguity_words = tokenize(ambiguity_text) if ambiguity_text else []
        innovation_words = tokenize(innovation_text) if innovation_text else []

        file_hash = sha256_file(pdf_path)
        documents.append(
            {
                "documentId": document_id,
                "companyKey": company_key,
                "year": year,
                "title": f"{item['companyName']}{year}年年度报告",
                "officialUrl": item["officialUrl"],
                "publicationDate": publication_date(item["officialUrl"]),
                "archiveStatus": "已归档",
                "parseStatus": "已解析" if mdna else "部分解析",
                "fileSha256": file_hash,
                "byteSize": pdf_path.stat().st_size,
                "pageCount": len(reader.pages),
                "sectionCoverage": {
                    "经营情况讨论与分析": bool(mdna),
                    "公司业务概要": bool(business),
                    "风险提示": bool(risk),
                    "业绩展望": bool(outlook),
                    "matchedHeadings": [x for x in [mdna_heading, business_heading, risk_heading, outlook_heading] if x],
                },
                "browserValidation": "Ego已确认交易所公开入口；归档文件使用公开镜像交叉下载",
            }
        )

        if disclosure_words:
            information_sufficiency = len(disclosure_words) / 10000
            numeric_count = len(NUMERIC_PATTERN.findall(disclosure_text)) + len(
                FINANCIAL_RATIO_PATTERN.findall(disclosure_text)
            )
            data_support = numeric_count / len(disclosure_words)
            observations.append(
                value_record(
                    company_key,
                    year,
                    "information_sufficiency",
                    information_sufficiency,
                    "已计算",
                    None,
                    document_id,
                    {"effectiveWordCount": len(disclosure_words)},
                )
            )
            observations.append(
                value_record(
                    company_key,
                    year,
                    "data_support_ratio",
                    data_support,
                    "已计算",
                    None,
                    document_id,
                    {"numericOccurrenceCount": numeric_count, "effectiveWordCount": len(disclosure_words)},
                )
            )
        else:
            information_sufficiency = None
            data_support = None
            reason = "未识别到经营情况讨论与分析章节"
            observations.append(value_record(company_key, year, "information_sufficiency", None, "缺失", reason, document_id))
            observations.append(value_record(company_key, year, "data_support_ratio", None, "缺失", reason, document_id))

        if ambiguity_words:
            uncertainty_count = sum(ambiguity_text.count(word) for word in UNCERTAINTY_WORDS)
            ambiguity = uncertainty_count / len(ambiguity_words)
            observations.append(
                value_record(
                    company_key,
                    year,
                    "risk_context_ambiguity",
                    ambiguity,
                    "已计算",
                    None,
                    document_id,
                    {"uncertaintyOccurrenceCount": uncertainty_count, "effectiveWordCount": len(ambiguity_words)},
                )
            )
        else:
            ambiguity = None
            observations.append(
                value_record(
                    company_key,
                    year,
                    "risk_context_ambiguity",
                    None,
                    "缺失",
                    "未识别到业绩展望或风险提示章节",
                    document_id,
                )
            )

        if information_sufficiency is not None and ambiguity is not None and data_support is not None:
            normalized_information = min(1.0, max(0.0, (information_sufficiency - 0.5) / 0.5))
            disclosure_quality = (normalized_information + (1 - ambiguity) + data_support) / 3
            observations.append(
                value_record(
                    company_key,
                    year,
                    "disclosure_quality",
                    disclosure_quality,
                    "已计算",
                    None,
                    document_id,
                    {"normalizedInformationSufficiency": round(normalized_information, 10)},
                )
            )
        else:
            observations.append(
                value_record(
                    company_key,
                    year,
                    "disclosure_quality",
                    None,
                    "缺失",
                    "三项子指标未全部满足计算条件",
                    document_id,
                )
            )

        talk = None
        if innovation_words:
            hits = innovation_hits(innovation_text)
            talk = hits / len(innovation_words) * 1000
            observations.append(
                value_record(
                    company_key,
                    year,
                    "innovation_talk_density",
                    talk,
                    "已计算",
                    None,
                    document_id,
                    {
                        "innovationOccurrenceCount": hits,
                        "effectiveWordCount": len(innovation_words),
                        "lexiconStatus": "程新生等（2022）完整词典口径",
                        "dictionaryRowCount": len(INNOVATION_DICTIONARY_ROWS),
                        "effectiveUniquePatternCount": len(INNOVATION_TERMS),
                    },
                )
            )
        else:
            observations.append(
                value_record(company_key, year, "innovation_talk_density", None, "缺失", "创新文本章节未满足提取条件", document_id)
            )

        patent_observation = ANNUAL_REPORT_INVENTION_APPLICATIONS.get((company_key, year))
        if patent_observation is None:
            patent_count = None
            observations.append(
                value_record(
                    company_key,
                    year,
                    "innovation_action_strength",
                    None,
                    "缺失",
                    "公开年报未明确披露当年发明专利申请数，国家知识产权局检索需登录，未以授权数、累计数或申请中存量替代",
                    document_id,
                )
            )
        else:
            patent_count, patent_basis = patent_observation
            observations.append(
                value_record(
                    company_key,
                    year,
                    "innovation_action_strength",
                    math.log1p(patent_count),
                    "已计算",
                    None,
                    document_id,
                    {
                        "annualInventionApplications": patent_count,
                        "verification": patent_basis,
                    },
                )
            )

        if talk is not None and patent_count is not None:
            divergence = math.log1p(talk) - math.log1p(patent_count)
            observations.append(
                value_record(
                    company_key,
                    year,
                    "innovation_divergence",
                    divergence,
                    "已计算",
                    None,
                    document_id,
                    {
                        "innovationTalkDensity": round(talk, 10),
                        "annualInventionApplications": patent_count,
                        "interpretation": "多言寡行" if divergence > 0 else "言行匹配或寡言多行",
                        "scheme": "方案二：企业自身言行对数差",
                    },
                )
            )
        else:
            missing_parts = []
            if talk is None:
                missing_parts.append("创新文本密度")
            if patent_count is None:
                missing_parts.append("当年发明专利申请数")
            observations.append(
                value_record(
                    company_key,
                    year,
                    "innovation_divergence",
                    None,
                    "缺失",
                    "、".join(missing_parts) + "缺失，方案二无法计算",
                    document_id,
                )
            )

    tone_audits = []
    for company_key, (_, _, _, years) in COMPANY_META.items():
        for year in years:
            tone_source = tone_sources.get((company_key, year))
            if tone_source and tone_source["file"].exists():
                payload = json.loads(tone_source["file"].read_text("utf-8"))
                records = payload.get("datas", [{}])[0].get("records", [])
                target_company_id = tone_source.get("targetCompanyId")
                answers = [
                    clean_answer_text(record.get("content", ""))
                    for record in records
                    if record.get("questionType") == 2
                    and record.get("isAnswered") is True
                    and record.get("content")
                    and (
                        target_company_id is None
                        or record.get("companyId") == target_company_id
                    )
                ]
                answers = [answer for answer in answers if answer]
                answer_text = "\n".join(answers)
                answer_words = tokenize(answer_text, remove_numeric=True)
                counts = Counter(answer_words)
                positive_count = sum(counts[word] for word in positive_words)
                negative_count = sum(counts[word] for word in negative_words)
                if answer_words:
                    positive_intensity = positive_count / len(answer_words)
                    negative_intensity = negative_count / len(answer_words)
                else:
                    positive_intensity = negative_intensity = None
                denominator = positive_count + negative_count
                tone = ((positive_count - negative_count) / denominator) if denominator else None
                source_document_id = f"performance-briefing:{company_key}:{year}"
                common_details = {
                    "answerCount": len(answers),
                    "effectiveWordCount": len(answer_words),
                    "positiveWordCount": positive_count,
                    "negativeWordCount": negative_count,
                    "sourceUrl": tone_source["url"],
                    "activityId": tone_source["activityId"],
                    "activityTitle": tone_source["activityTitle"],
                    "matchingMethod": "用户提供的姜富伟等（2020）中文金融情感词典词频法",
                }
                observations.extend(
                    [
                        value_record(company_key, year, "positive_tone_intensity", positive_intensity, "已计算", None, source_document_id, common_details),
                        value_record(company_key, year, "negative_tone_intensity", negative_intensity, "已计算", None, source_document_id, common_details),
                        value_record(
                            company_key,
                            year,
                            "manager_net_positive_tone",
                            tone,
                            "已计算" if tone is not None else "缺失",
                            None if tone is not None else "正面词与负面词均为零",
                            source_document_id,
                            {**common_details, "riskLabel": risk_label(tone)},
                        ),
                    ]
                )
                tone_audits.append(
                    {
                        "companyKey": company_key,
                        "year": year,
                        "formalMethod": "姜富伟等中文金融情感词典词频法",
                        "sourceUrl": tone_source["url"],
                        "answerCount": len(answers),
                        "dictionaryReview": "已完成",
                        "modelReview": "不适用",
                        "modelReviewReason": "用户指定指标三不使用大模型，正式结果仅使用所提供情感词典",
                    }
                )
            else:
                reason = tone_exclusions.get(
                    (company_key, year),
                    "未定位到可公开验证且满足排除规则的上证路演中心年度业绩说明会管理层回答全集",
                )
                for metric_key in (
                    "positive_tone_intensity",
                    "negative_tone_intensity",
                    "manager_net_positive_tone",
                ):
                    observations.append(value_record(company_key, year, metric_key, None, "缺失", reason, None))

    observation_index = {(x["companyKey"], x["year"], x["metricKey"]): x for x in observations}
    for company_key, (_, _, _, years) in COMPANY_META.items():
        for metric in METRICS:
            previous = None
            for year in years:
                record = observation_index.get((company_key, year, metric["metricKey"]))
                if not record:
                    continue
                current = record["value"]
                if current is not None and previous is not None:
                    record["changeRate"] = round((current - previous) / (abs(previous) + 1e-6), 10)
                previous = current

    apply_risk_scores(observations)

    documents.sort(key=lambda item: (item["companyKey"], item["year"]))
    observations.sort(key=lambda item: (item["metricKey"], item["companyKey"], item["year"]))
    method_hash = sha256_file(METHOD_DOCUMENT_PATH) if METHOD_DOCUMENT_PATH.exists() else None
    snapshot = {
        "schemaVersion": SCHEMA_VERSION,
        "dataVersion": METHOD_VERSION,
        "asOfDate": AS_OF_DATE,
        "sourceMode": "snapshot",
        "methodVersion": {
            "methodVersion": METHOD_VERSION,
            "name": "叙事风险维度测度（修改版）",
            "effectiveDate": AS_OF_DATE,
            "sourceDocumentSha256": method_hash,
            "innovationLexiconStatus": "程新生等（2022）完整词典口径",
            "innovationLexiconSize": len(INNOVATION_DICTIONARY_ROWS),
            "innovationLexiconUniqueSize": len(INNOVATION_TERMS),
            "innovationLexiconCategoryCounts": dict(INNOVATION_CATEGORY_COUNTS),
            "innovationLexiconSha256": sha256_file(INNOVATION_LEXICON_PATH),
            "stopwordListSha256": sha256_file(STOPWORDS_PATH),
            "sentimentDictionaryName": "姜富伟等中文金融情感词典",
            "sentimentDictionarySha256": sentiment_hash,
            "sentimentDictionarySource": "https://github.com/MengLingchao/Chinese_financial_sentiment_dictionary",
            "peerBenchmarkStatus": "不适用：叙事夸大度采用方案二，不需要同行业基准",
            "notes": [
                "此前指标文件、旧词库、旧公式和旧结果不参与新版趋势计算或展示。",
                "综合披露质量按已确认解释使用三项等权，不构造跨维度总分。",
                "创新词典按用户提供工作簿读取：692条记录、691个唯一匹配词项；专利一词跨类别重复时不重复计数。",
                "叙事夸大度采用方案二：创新文本密度与当年发明专利申请数分别加一取自然对数后作差。",
                "管理者语调只使用上证路演中心管理层回答和用户提供的姜富伟等（2020）中文金融情感词典，不使用大模型。",
                "0至100风险分按风险方向在当前五家上市样本窗口内做极差映射；样本最小值、最大值和有效观测数随方法元数据公开。",
            ],
        },
        "companies": companies,
        "methodology": METRICS,
        "documents": documents,
        "observations": observations,
        "peerBenchmarks": [],
        "toneAudits": tone_audits,
        "audit": {
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "targetReportCount": 21,
            "archivedReportCount": sum(1 for item in documents if item["archiveStatus"] == "已归档"),
            "parsedReportCount": sum(1 for item in documents if item["parseStatus"] == "已解析"),
            "partialReportCount": sum(1 for item in documents if item["parseStatus"] == "部分解析"),
            "toneYearCount": len(tone_audits),
            "peerBenchmarkYearCount": 0,
            "missingObservationCount": sum(1 for item in observations if item["value"] is None),
            "calculatedObservationCount": sum(1 for item in observations if item["value"] is not None),
            "publicPayloadContainsFullText": False,
            "publicPayloadContainsPrivatePath": False,
        },
    }
    SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_PATH.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n", "utf-8")
    print(
        json.dumps(
            {
                "snapshot": str(SNAPSHOT_PATH.relative_to(ROOT)),
                "documents": len(documents),
                "archived": snapshot["audit"]["archivedReportCount"],
                "calculated": snapshot["audit"]["calculatedObservationCount"],
                "missing": snapshot["audit"]["missingObservationCount"],
                "toneYears": len(tone_audits),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
